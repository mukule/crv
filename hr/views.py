from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import user_passes_test
from applicants.models import *
from django.core.paginator import *
from django.db.models import Q
from django.db.models import Count
from applicants.forms import *
from django.contrib import messages
import csv
from django.http import HttpResponse
from applicants.models import *
import openpyxl
from openpyxl.styles import Alignment
import os
from django.http import Http404
import shutil
from django.core.mail import send_mail
# Create your views here.


def is_superuser(user):
    return user.is_superuser


@user_passes_test(is_superuser)
def index(request):
    return render(request, 'hr/index.html')


@user_passes_test(is_superuser)
def regs(request):
    # Get the search query from the request GET parameters
    search_query = request.GET.get('search', '')
    show_all = request.GET.get('show_all')

    # Start with all users (non-superusers)
    users = CustomUser.objects.filter(is_superuser=False)
    users = users.order_by('-date_joined')


    if search_query:
        # Filter by username, first name, and last name
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )

    if show_all:
        # If the "Show All" button is clicked, remove the search filter
        users = CustomUser.objects.filter(is_superuser=False)

    user_count = users.count()

    # Pagination
    paginator = Paginator(users, 100)  # Show 100 users per page
    page = request.GET.get('page')
    users = paginator.get_page(page)

    return render(request, 'hr/regs.html', {'users': users, 'q': search_query, 'user_count': user_count})


@user_passes_test(is_superuser)
def vac(request):
    # Get the values of query parameters (if they exist)
    job_name = request.GET.get('job_name')
    vacancy_type = request.GET.get('vacancy_type')
    job_ref = request.GET.get('job_ref')
    show_all = request.GET.get('show_all')

    # Start with an unfiltered queryset
    vacancies = Vacancy.objects.all().order_by('-date_created')
    if show_all:
        vacancies = Vacancy.objects.all().order_by('-date_created')

    # Apply filters based on query parameters
    if job_name:
        vacancies = vacancies.filter(
            job_name__icontains=job_name)  # Case-insensitive filter

    if vacancy_type:
        # Case-insensitive filter on vacancy type name
        vacancies = vacancies.filter(
            vacancy_type__name__icontains=vacancy_type)

    if job_ref:
        vacancies = vacancies.filter(
            job_ref__icontains=job_ref)  # Case-insensitive filter

    vacancies = vacancies.annotate(application_count=Count('jobapplication'))

    vac_count = vacancies.count()

    # Paginate the list of vacancies by 10 items per page
    paginator = Paginator(vacancies, 10)
    page = request.GET.get('page')

    try:
        vacancies = paginator.page(page)
    except PageNotAnInteger:
        vacancies = paginator.page(1)
    except EmptyPage:
        vacancies = paginator.page(paginator.num_pages)

    return render(request, 'hr/vac.html', {'vacancies': vacancies, 'vac_count': vac_count})


@user_passes_test(is_superuser)
def create_vac(request):
    if request.method == 'POST':
        form = VacancyForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Vacancy created successfully.')
            return redirect('hr:vac')
        else:
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")
            error_message = "Vacancy creation failed. Please check the form. Errors: " + \
                ', '.join(error_messages)
            messages.error(request, error_message)
    else:
        form = VacancyForm()

    return render(request, 'hr/create_vac.html', {'form': form})


@user_passes_test(is_superuser)
def edit_vac(request, vacancy_id):
    vacancy = get_object_or_404(Vacancy, id=vacancy_id)

    if request.method == 'POST':
        form = VacancyForm(request.POST, request.FILES, instance=vacancy)
        if form.is_valid():
            form.save()
            messages.success(request, 'Vacancy updated successfully.')
            return redirect('hr:vac')
        else:
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")
            error_message = "Vacancy update failed. Please check the form. Errors: " + \
                ', '.join(error_messages)
            messages.error(request, error_message)
    else:
        form = VacancyForm(instance=vacancy)

    return render(request, 'hr/edit_vacancy.html', {'form': form, 'vacancy': vacancy})


@user_passes_test(is_superuser)
def vac_ap(request):
    # Query to retrieve all vacancies along with the number of applications for each,
    job_name = request.GET.get('job_name')
    vacancy_type = request.GET.get('vacancy_type')
    job_ref = request.GET.get('job_ref')
    show_all = request.GET.get('show_all')
    # ordered by the latest vacancies first
    vacancies = Vacancy.objects.annotate(
        application_count=Count('jobapplication')).order_by('-date_open')

    if show_all:
        vacancies = Vacancy.objects.all().order_by('-date_created')

    # Apply filters based on query parameters
    if job_name:
        vacancies = vacancies.filter(
            job_name__icontains=job_name)  # Case-insensitive filter

    if vacancy_type:
        # Case-insensitive filter on vacancy type name
        vacancies = vacancies.filter(
            vacancy_type__name__icontains=vacancy_type)

    if job_ref:
        vacancies = vacancies.filter(
            job_ref__icontains=job_ref)  # Case-insensitive filter

    vacancies = vacancies.annotate(application_count=Count('jobapplication'))

    vac_count = vacancies.count()

    # Paginate the list of vacancies by 10 items per page
    paginator = Paginator(vacancies, 10)
    page = request.GET.get('page')

    try:
        vacancies = paginator.page(page)
    except PageNotAnInteger:
        vacancies = paginator.page(1)
    except EmptyPage:
        vacancies = paginator.page(paginator.num_pages)

    context = {
        'vacancies': vacancies,
        'vac_count': vac_count
    }

    return render(request, 'hr/ap.html', context)


@user_passes_test(is_superuser)
def vac_ap_detail(request, vacancy_id):
    vacancy = get_object_or_404(Vacancy, pk=vacancy_id)
    applications = JobApplication.objects.filter(vacancy=vacancy)

    # Search functionality
    search_query = request.GET.get('search')
    gender_filter = request.GET.get('gender')
    disability_filter = request.GET.get('disability')
    qualified_status_filter = request.GET.get('qualified_status')
    marital_status_filter = request.GET.get('marital_status')
    shortlisted_filter = request.GET.get('shortlisted')
    show_all = request.GET.get('show_all')

    if search_query:
        # Filter by username, first name, last name, gender, and disability within job applications
        applications = applications.filter(
            Q(user__username__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query)
        )

    if gender_filter:
        applications = applications.filter(user__gender=gender_filter)

    if disability_filter:
        applications = applications.filter(
            user__is_person_with_disability=disability_filter)

    if qualified_status_filter is not None:
        applications = applications.filter(
            is_qualified=qualified_status_filter)

    if marital_status_filter:
        applications = applications.filter(
            user__marital_status=marital_status_filter)

    if shortlisted_filter is not None:
        applications = applications.filter(
            shortlisted=shortlisted_filter)

    if show_all:
        # If the "Show All" button is clicked, remove the search filter
        applications = JobApplication.objects.filter(vacancy=vacancy)

    # Check if the user wants to export to Excel
        # Check if the user wants to export to Excel
    export_excel = request.GET.get('export_excel')
    if export_excel:
        # Create a list of dictionaries representing the data you want to export
        data = []
        for application in applications:
            user = application.user
            resume = user.Resume
            academic_details = user.AcademicDetails.all()[:3]
            relevant_courses = user.Relevantcourse.all()[:3]
            employment_history = user.EmploymentHistory.all()[:3]
            referees = user.Referee.all()[:3]

            application_data = {
                'Username/ID NO.': user.username,
                'Full Name': f"{user.first_name} {user.last_name}",
                'Date Applied': application.application_date.strftime("%Y-%m-%d %H:%M:%S"),
                'Gender': user.gender,
                'Disability': 'Yes' if user.is_person_with_disability == 'Y' else 'No',
                'Qualification Status': 'Qualified' if application.is_qualified else 'Not Qualified',
                'Shortlisted': 'Yes' if application.shortlisted else 'No',
                'Education': '',
                'Relevant Course': '',
                'Employment History': '',
                'Referee': '',
            }

            if academic_details:
                # Handle Academic Details data
                academic_data = ''
                for academic_detail in academic_details:
                    institution_name = f"Institution Name: {academic_detail.institution_name}"
                    admission_number = f"Admission Number: {academic_detail.admission_number}"
                    start_year = f"Start Year: {academic_detail.start_year.strftime('%Y')}"
                    end_year = f"End Year: {academic_detail.end_year.strftime('%Y')}" if academic_detail.end_year else "End Year: In Progress"
                    graduation_year = f"Graduation Year: {academic_detail.graduation_year.strftime('%Y')}" if academic_detail.graduation_year else "Graduation Year: In Progress"
                    academic_data += f"{institution_name}\n{admission_number}\n{start_year}\n{end_year}\n{graduation_year}\n\n"

                application_data['Education'] = academic_data

            if relevant_courses:
                # Handle Relevant Course data
                relevant_course_data = ''
                for relevant_course in relevant_courses:
                    course_name = f"Course Name: {relevant_course.course_name}"
                    institution = f"Institution: {relevant_course.institution}"
                    certification = f"Certification: {relevant_course.certification}"
                    start_date = f"Start Date: {relevant_course.start_date.strftime('%Y-%m-%d')}" if relevant_course.start_date else "Start Date: Not specified"
                    completion_date = f"Completion Date: {relevant_course.completion_date.strftime('%Y-%m-%d')}" if relevant_course.completion_date else "Completion Date: In Progress"
                    relevant_course_data += f"{course_name}\n{institution}\n{certification}\n{start_date}\n{completion_date}\n\n"

                application_data['Relevant Course'] = relevant_course_data

            if employment_history:
                # Handle Employment History data
                employment_data = ''
                for employment_detail in employment_history:
                    company_name = f"Company Name: {employment_detail.company_name}"
                    position = f"Position: {employment_detail.position}"
                    position_description = f"Position Description: {employment_detail.position_description}"
                    start_date = f"Start Date: {employment_detail.start_date.strftime('%Y-%m-%d')}" if employment_detail.start_date else "Start Date: Not specified"
                    end_date = f"End Date: {employment_detail.end_date.strftime('%Y-%m-%d')}" if employment_detail.end_date else "End Date: In Progress"
                    employment_data += f"{company_name}\n{position}\n{position_description}\n{start_date}\n{end_date}\n\n"

                application_data['Employment History'] = employment_data

            if referees:
                # Handle Referee data
                referee_data = ''
                for referee_detail in referees:
                    referee_name = f"Referee Name: {referee_detail.name}"
                    occupation = f"Occupation: {referee_detail.occupation}"
                    organization = f"Organization: {referee_detail.organization}"
                    relationship_period = f"Relationship Period: {referee_detail.relationship_period}"
                    email = f"Email: {referee_detail.email}" if referee_detail.email else "Email: Not specified"
                    phone = f"Phone: {referee_detail.phone}"
                    referee_data += f"{referee_name}\n{occupation}\n{organization}\n{relationship_period}\n{email}\n{phone}\n\n"

                application_data['Referee'] = referee_data

            data.append(application_data)

        # Define column headers for Excel export
        headers = [
            'Username/ID NO.', 'Full Name', 'Date Applied', 'Gender', 'Disability', 'Qualification Status', 'Shortlisted', 'Education', 'Relevant Course', 'Employment History', 'Referee'
            # Add more headers as needed
        ]

        # Create a workbook and add a worksheet
        wb = openpyxl.Workbook()
        ws = wb.active

        # Write title
        title = f"{vacancy.job_name} - {vacancy.job_ref} Applications"
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.alignment = Alignment(horizontal='center')
        title_cell.font = openpyxl.styles.Font(size=14, bold=True)

        # Write headers
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=2, column=col_num, value=header)
            # Adjust the width of the 'Education' column
            if header == 'Education':
                ws.column_dimensions[openpyxl.utils.get_column_letter(
                    col_num)].width = 30

            if header == 'Employment History':
                ws.column_dimensions[openpyxl.utils.get_column_letter(
                    col_num)].width = 30

            if header == 'Referee':
                ws.column_dimensions[openpyxl.utils.get_column_letter(
                    col_num)].width = 30

            if header == 'Relevant Course':
                ws.column_dimensions[openpyxl.utils.get_column_letter(
                    col_num)].width = 30
            # Adjust the width of other columns as needed

        # Write data
        for row_num, application_data in enumerate(data, 3):
            for col_num, value in enumerate(application_data.values(), 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                # Adjust alignment for 'Education' column
                if headers[col_num - 1] == 'Education':
                    cell.alignment = Alignment(wrap_text=True)

                if headers[col_num - 1] == 'Referee':
                    cell.alignment = Alignment(wrap_text=True)

                if headers[col_num - 1] == 'Relevant Course':
                    cell.alignment = Alignment(wrap_text=True)

                if headers[col_num - 1] == 'Employment History':
                    cell.alignment = Alignment(wrap_text=True)

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{vacancy.job_ref}.xlsx"'

        # Save the workbook to the response
        wb.save(response)

        return response

    vac_count = applications.count()
    # Pagination
    paginator = Paginator(applications, 100)  # Show 20 applications per page
    page = request.GET.get('page')
    applications = paginator.get_page(page)
    # Get gender, disability, and marital status choices
    gender_choices = CustomUser.GENDER_CHOICES
    disability_choices = CustomUser.DISABILITY_CHOICES
    marital_status_choices = CustomUser.MARITAL_STATUS_CHOICES

    context = {
        'vacancy': vacancy,
        'applications': applications,
        'search_query': search_query,
        'vac_count': vac_count,
        'gender_choices': gender_choices,
        'disability_choices': disability_choices,
        'marital_status_choices': marital_status_choices,
    }

    return render(request, 'hr/vac_ap_detail.html', context)


@user_passes_test(is_superuser)
def shortlist(request, application_id):
    job_application = get_object_or_404(JobApplication, pk=application_id)

    # Toggle the shortlisted status
    job_application.shortlisted = not job_application.shortlisted
    job_application.save()

    # Redirect back to the vacancy detail page or any other page you prefer
    return redirect('hr:vac_ap_detail', vacancy_id=job_application.vacancy.id)


@user_passes_test(is_superuser)
def zip_and_download_documents(user):
    # Create a temporary directory to store the documents
    temp_dir = 'temp_documents'
    os.makedirs(temp_dir, exist_ok=True)

    # Copy the documents to the temporary directory
    for document in user.Document.all():
        file_path = document.document.path
        shutil.copy(file_path, temp_dir)

    # Create a zip file
    zip_file_path = f"{temp_dir}.zip"
    shutil.make_archive(temp_dir, 'zip', temp_dir)

    # Remove the temporary directory
    shutil.rmtree(temp_dir)

    return zip_file_path


@user_passes_test(is_superuser)
def resume(request, user_id):
    user = get_object_or_404(CustomUser, pk=user_id)

    academic_details = user.AcademicDetails.all()[:3]
    relevant_courses = user.Relevantcourse.all()[:3]
    employment_history = user.EmploymentHistory.all()[:3]
    referees = user.Referee.all()[:3]
    documents = user.Document.all()

    print(documents)
    if request.method == 'POST' and 'download_docs' in request.POST:
        try:
            zip_file_path = zip_and_download_documents(user)
            with open(zip_file_path, 'rb') as zip_file:
                response = HttpResponse(
                    zip_file.read(), content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename=documents.zip'
                return response
        except FileNotFoundError:
            messages.error(
                request, "An error occurred while processing the download. Please try again later.")
            return redirect('hr:resume', user_id)

    context = {
        'user': user,
        'academic': academic_details,
        'courses': relevant_courses,
        'experience': employment_history,
        'ref': referees,
        'docs': documents
    }

    return render(request, 'hr/resume.html', context)


@user_passes_test(is_superuser)
def update_response(request, application_id):
    job_application = get_object_or_404(JobApplication, pk=application_id)

    if request.method == 'POST':
        form = JobApplicationResponseForm(
            request.POST, instance=job_application)
        if form.is_valid():
            form.save()

            # Check if the send_email checkbox is selected
            if form.cleaned_data['send_email']:
                # Customize the email subject and content as needed
                vacancy_name = job_application.vacancy.job_name
                vacancy_ref = job_application.vacancy.job_ref
                email_subject = f'Application Feedback for {vacancy_name} (Ref: {vacancy_ref})'
                email_message = f"Here is your update: {job_application.response}"

                recipient_email = job_application.user.email

                # Send the email
                send_mail(email_subject, email_message,
                          'recruitment@crvwwda.go.ke', [recipient_email])

                # Add a success message for sending the email
                messages.success(
                    request, 'Email notification sent successfully.')

            # Add a success message for updating the response
            messages.success(request, 'Feedback added successfully.')

            # Redirect to the vacancy application detail page
            return redirect('hr:vac_ap_detail', job_application.vacancy.id)
    else:
        form = JobApplicationResponseForm(instance=job_application)

    return render(request, 'hr/feedback.html', {'form': form, 'app': job_application})
